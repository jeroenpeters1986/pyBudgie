import openpyxl
from datetime import datetime

from django.contrib import admin, messages
from django.contrib.admin import SimpleListFilter
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _

from budgie_user.mixins import BudgieUserMixin
from .forms import BirdForm
from .mixins import AdminExportCsvMixin, AdminExportAllCsvMixin
from .models import Bird, Breeder, ColorProperty, BirdProxy, BirdPhoto


class BirdPhotoInline(admin.StackedInline):
    model = BirdPhoto
    extra = 1


class BirdFatherFilter(SimpleListFilter):
    title = _("father")
    parameter_name = "father"

    def lookups(self, request, model_admin):
        males = Bird.objects.filter(gender=Bird.Gender.MALE)
        return [(b.pk, str(b)) for b in males]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(father__pk=self.value())
        return queryset


class BirdMotherFilter(SimpleListFilter):
    title = _("mother")
    parameter_name = "mother"

    def lookups(self, request, model_admin):
        males = Bird.objects.filter(gender=Bird.Gender.FEMALE)
        return [(b.pk, str(b)) for b in males]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(mother__pk=self.value())
        return queryset


@admin.register(Bird)
class BirdAdmin(
    BudgieUserMixin, admin.ModelAdmin, AdminExportCsvMixin, AdminExportAllCsvMixin
):
    form = BirdForm

    inlines = [BirdPhotoInline]

    list_display = [
        "ring_number",
        "gender",
        "color",
        "color_props",
        "split_props",
        "date_of_birth",
        "current_age",
        "image_tag",
        "family_tree",
    ]
    list_filter = [
        "gender",
        "color",
        "color_property",
        "split_property",
        BirdFatherFilter,
        BirdMotherFilter,
        "is_owned",
        "is_for_sale",
    ]
    search_fields = ["ring_number", "gender"]
    search_help_text = _(
        "Search for ring numbers. "
        "You can also narrow down your search using the filters on the right."
    )
    date_hierarchy = "date_of_birth"
    ordering = ["ring_number"]
    show_facets = admin.ShowFacets.ALWAYS

    fieldsets = (
        (
            _("General information"),
            {
                "fields": (
                    "user",
                    "ring_number",
                    "gender",
                    "date_of_birth",
                    "date_of_death",
                )
            },
        ),
        (
            _("Color properties"),
            {
                "fields": ("color", "color_property", "split_property"),
            },
        ),
        (
            _("Family tree"),
            {
                "fields": ("father", "mother", "breeder"),
            },
        ),
        (
            _("Other information"),
            {
                "fields": ("owner", "is_owned", "is_for_sale"),
            },
        ),
        (None, {"fields": ("photo", "notes")}),
    )

    autocomplete_fields = ["father", "mother", "breeder", "owner"]
    save_on_top = True
    save_as = True
    actions = [
        "export_as_csv",
        "export_all_as_csv",
        "mark_as_owned",
        "mark_as_for_sale",
    ]

    change_list_template = "budgie_bird/admin/bird_changelist.html"

    def get_urls(self):
        """Extend the urls of the Django admin with new views"""
        urls = super().get_urls()
        additional_bird_admin_urls = [
            path(
                "<path:object_id>/family_tree/",
                self.admin_site.admin_view(self.family_tree_view, cacheable=True),
                name="budgie_bird_bird_familytree",
            )
        ]
        return additional_bird_admin_urls + urls

    def get_readonly_fields(self, request, obj=None):
        """
        This makes sure the admin doesn't give away the other usernames to non-admins.
        """
        if not request.user.is_superuser:
            return ["user"]
        return {}

    def image_tag(self, obj):
        """Render the image tag of the birds photo"""
        return mark_safe(
            '<img src="{}" height="48" class="birdpreview" />'.format(obj.photo.url)
        )

    image_tag.short_description = _("Photo")

    def current_age(self, obj):
        """Calculate the age of the bird"""

        if not obj.date_of_birth:
            return "-"

        today = obj.date_of_birth.today()
        years = (
            today.year
            - obj.date_of_birth.year
            - (
                (today.month, today.day)
                < (obj.date_of_birth.month, obj.date_of_birth.day)
            )
        )
        months = today.month - obj.date_of_birth.month
        if months < 0:
            months += 12

        return mark_safe(
            "<abbr title='{} {}, {} {}'>{}{}, {}{}</abbr>".format(
                years,
                _("years"),
                months,
                _("months"),
                years,
                _("years")[0],
                months,
                _("months")[0],
            )
        )
        # Too long?  return "{} {}, {} {}".format(years, _("years"), months, _("months"))

    current_age.short_description = _("Age")

    def family_tree(self, obj):
        return mark_safe(
            '<a class="grp-button" href="{}">{}</a>'.format(
                reverse("admin:budgie_bird_bird_familytree", args=[obj.pk]), _("View")
            )
        )

    family_tree.short_description = _("Family tree")

    def color_props(self, obj):
        return obj.color_props()

    color_props.short_description = _("Color properties")

    def split_props(self, obj):
        return obj.split_props()

    split_props.short_description = _("Split properties")

    def mark_as_owned(self, request, queryset):
        queryset.update(is_owned=True)
        messages.add_message(
            request, messages.SUCCESS, _("Selected birds are marked as owned")
        )

    mark_as_owned.short_description = _("Mark as owned")

    def mark_as_for_sale(self, request, queryset):
        queryset.update(is_for_sale=True)
        messages.add_message(
            request, messages.SUCCESS, _("Selected birds are marked as for sale")
        )

    mark_as_for_sale.short_description = _("Mark as for sale")

    def convert_bird_to_treantjs_data(self, bird):
        tree_data = {
            "HTMLclass": "pyBudgie_{}".format(bird.gender),
            "text": {
                "name": bird.ring_number,
                "desc": "{}: {}".format(_("Date of birth"), bird.date_of_birth or ""),
                "contact": bird.descriptive_color(),
            },
            "image": bird.photo.url,
        }

        children = []
        if bird.father:
            children.append(self.convert_bird_to_treantjs_data(bird.father))
        if bird.mother:
            children.append(self.convert_bird_to_treantjs_data(bird.mother))

        if children:
            tree_data["children"] = children

        return tree_data

    def family_tree_view(self, request, *args, **kwargs):
        """Custom admin view to show the family tree"""
        try:
            bird = Bird.objects.get(pk=kwargs["object_id"])
        except Bird.DoesNotExist:
            messages.add_message(request, messages.ERROR, _("That bird does not exist"))
            return redirect(reverse("admin:budgie_bird_bird_changelist"))

        context = dict(
            self.admin_site.each_context(request),  # Common admin things
            bird=bird,
            family_tree_data=mark_safe(self.convert_bird_to_treantjs_data(bird)),
        )
        return TemplateResponse(
            request, "budgie_bird/admin/bird_familytree.html", context
        )


@admin.register(Breeder)
class BreederAdmin(BudgieUserMixin, admin.ModelAdmin):
    search_fields = ["first_name", "last_name", "breeding_reg_nr", "notes", "address"]
    list_display = ["display_name", "breeding_reg_nr", "phone_number"]

    def display_name(self, obj):
        # We wouldn't need this if it wasn't for translations...
        return obj.display_name()

    display_name.short_description = _("Full name")


@admin.register(ColorProperty)
class ColorPropertyAdmin(BudgieUserMixin, admin.ModelAdmin):
    search_fields = ["color_name"]
    list_display = ["color_name", "rank"]
    list_editable = ["rank"]


@admin.register(BirdProxy)
class ExportBirdAdmin(admin.ModelAdmin):
    def export_view(self, request):
        if request.method != "POST":
            # Display the download-button page
            opts = self.model._meta
            app_label = opts.app_label
            object_name = opts.verbose_name
            object_url = reverse(
                "admin:%s_%s_changelist" % (app_label, opts.model_name)
            )
            context = dict(
                self.admin_site.each_context(request),
                title="Download %s" % object_name,
                object_name=object_name,
                object_url=object_url,
            )
            return TemplateResponse(
                request, "budgie_bird/admin/bird_export.html", context
            )

        queryset = self.model.objects.filter(user=request.user)

        # Maak een nieuw Excel-werkboek en werkblad
        excel_workbook = openpyxl.Workbook()
        excel_sheet = excel_workbook.active

        # Definieer de veldnamen/headers
        headers = [
            "Ringnummer",
            "Kleur",
            "Kleurcategorie",
            "Kleurslagen",
            "Split voor",
            "Vader",
            "Moeder",
            "Geboren",
            "Kweker",
            "Eigenaar",
            "Geslacht",
            "In bezit",
            "Te Koop",
            "Notities",
        ]

        # Headers
        for col_num, header in enumerate(headers, 1):
            excel_sheet.cell(row=1, column=col_num, value=header)

        # Bird data
        for row_num, bird in enumerate(queryset, 2):
            column_num = 0

            # Ring number
            column_num += 1
            excel_sheet.cell(row=row_num, column=column_num, value=bird.ring_number)

            # Color name full-out
            column_num += 1
            excel_sheet.cell(
                row=row_num, column=column_num, value=bird.descriptive_color()
            )

            # Color (catalog) category
            column_num += 1
            excel_sheet.cell(
                row=row_num, column=column_num, value=bird.get_color_display()
            )

            # Color properties
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=bird.color_props(),
            )

            # Split properties
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=bird.split_props(),
            )

            # Father
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=self.get_model_string_or_empty(bird.father),
            )

            # Mother
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=self.get_model_string_or_empty(bird.mother),
            )

            # Date of birth
            column_num += 1
            excel_sheet.cell(row=row_num, column=column_num, value=bird.date_of_birth)

            # Breeder
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=self.get_model_string_or_empty(bird.breeder),
            )

            # Owner
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=self.get_model_string_or_empty(bird.owner),
            )

            # Sex
            column_num += 1
            excel_sheet.cell(
                row=row_num, column=column_num, value=bird.get_gender_display()
            )

            # Is this bird owned by the account holder?
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=str(self.get_value_yes_no(bird.is_owned)),
            )

            # Is this bird for sale?
            column_num += 1
            excel_sheet.cell(
                row=row_num,
                column=column_num,
                value=str(self.get_value_yes_no(bird.is_for_sale)),
            )

            # Notes about the bird
            column_num += 1
            excel_sheet.cell(row=row_num, column=column_num, value=bird.notes)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            "attachment; filename={filename}-{timestamp}.xlsx".format(
                filename=_("bird_export"),
                timestamp=str(datetime.now().strftime("%Y-%m-%d")),
            )
        )

        # Save this Excel workbook to the response
        excel_workbook.save(response)

        return response

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("export_file/", self.admin_site.admin_view(self.export_view)),
        ]
        return my_urls + urls

    def get_model_string_or_empty(self, bird):
        return str(bird) if bird else ""

    def get_value_yes_no(self, bird_boolean):
        return _("Yes") if bird_boolean else _("No")
